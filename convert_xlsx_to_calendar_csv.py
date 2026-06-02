#!/usr/bin/env python3
"""
convert_xlsx_to_calendar_csv.py

Converts the exported Google Sheets CSV (or xlsx) into the format
expected by the GitHub calendar generator.

Usage:
    python convert_xlsx_to_calendar_csv.py [input_file] [output_file]

    input_file:  path to the exported CSV or xlsx from Google Sheets
                 (default: "26-27_Admin_Calendar_Events_CTE-LL_Version.csv")
    output_file: path to write the GitHub-ready CSV
                 (default: "CTE-LL_Calendar_Events_2026-27.csv")

Handles:
    - CSV or xlsx input
    - Combines Date + End Date into YYYY-MM-DD:YYYY-MM-DD range format
    - Extracts team emoji from Event Name into Team column
    - Converts "● Confirmed" / "◆ Hold" status to "Confirmed" / "Hold"
    - Drops unused columns (Region/Division, Add to Master, Add to CTE-LL)
    - Skips TBD rows (no parseable date)
"""

import sys
import os
import re
import csv
from datetime import datetime

# ── Team emoji → identifier ───────────────────────────────────────────────────
# Order matters: check longer/more-specific patterns first
TEAM_EMOJIS = ["📰🐦▶️", "🧰", "🐦", "▶️", "📰", "📗", "💰", "📧", "🏫"]

def extract_team(event_name):
    """Return the first team emoji found in the event name, or empty string."""
    for emoji in TEAM_EMOJIS:
        if emoji in event_name:
            return emoji
    return ""

# ── Date parsing ──────────────────────────────────────────────────────────────
DATE_FORMATS = [
    "%A, %m/%d/%y",   # "Wednesday, 7/8/26"
    "%A, %m/%d/%Y",   # "Wednesday, 7/8/2026"
    "%m/%d/%Y",       # "7/8/2026"
    "%m/%d/%y",       # "7/8/26"
    "%Y-%m-%d",       # already ISO
    "%m/%d/%Y %H:%M:%S",  # with time (xlsx export artifact)
]

def parse_date(raw):
    """Parse a date string into YYYY-MM-DD, or None if unparseable."""
    if not raw:
        return None
    raw = str(raw).strip()
    # Already ISO
    if re.match(r'^\d{4}-\d{2}-\d{2}$', raw):
        return raw
    # Strip time component if present
    raw_clean = re.sub(r'\s+\d{1,2}:\d{2}(:\d{2})?(\s*(AM|PM))?$', '', raw, flags=re.IGNORECASE).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(raw_clean, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None

def format_date_field(start_raw, end_raw):
    """Return ISO date or ISO:ISO range, or None if start is unparseable."""
    start = parse_date(start_raw)
    if not start:
        return None
    end = parse_date(end_raw)
    if end and end != start:
        return f"{start}:{end}"
    return start

def normalize_status(raw):
    """Strip bullet/diamond prefix and return Confirmed, Hold, or empty."""
    s = (raw or "").strip()
    if "Confirmed" in s or s == "●":
        return "Confirmed"
    if "Hold" in s or s == "◆":
        return "Hold"
    return ""

# ── Read input ────────────────────────────────────────────────────────────────
def read_input(path):
    """Return list of row dicts from a CSV or xlsx file."""
    ext = os.path.splitext(path)[1].lower()

    if ext in (".xlsx", ".xlsm", ".xls"):
        try:
            import openpyxl
        except ImportError:
            print("❌ openpyxl not installed. Run: pip install openpyxl")
            sys.exit(1)
        wb = openpyxl.load_workbook(path, data_only=True)
        # Use first sheet, or "26-27 Events" if present
        ws = wb["26-27 Events"] if "26-27 Events" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        # Find header row (first row containing "Event Name")
        header_row = None
        for i, row in enumerate(rows):
            if any("Event Name" in str(c) for c in row if c):
                header_row = i
                break
        if header_row is None:
            print("❌ Could not find header row in xlsx")
            sys.exit(1)
        headers = [str(c).strip() if c else "" for c in rows[header_row]]
        return [dict(zip(headers, row)) for row in rows[header_row + 1:]]

    else:  # CSV
        with open(path, newline="", encoding="utf-8-sig") as f:
            # Sniff for header row
            content = f.read()
        lines = content.splitlines()
        header_idx = 0
        for i, line in enumerate(lines):
            if "Event Name" in line:
                header_idx = i
                break
        reader = csv.DictReader(lines[header_idx:])
        return list(reader)

# ── Column name aliases ───────────────────────────────────────────────────────
def get(row, *keys):
    """Try multiple possible column name variants, return first match."""
    for k in keys:
        for rk in row:
            if rk and k.lower().strip() in rk.lower().strip():
                v = row[rk]
                return str(v).strip() if v is not None else ""
    return ""

# ── Main ──────────────────────────────────────────────────────────────────────
def convert(input_path, output_path):
    rows = read_input(input_path)

    OUT_FIELDS = ["Date", "Status", "Team", "Event Name",
                  "Audience", "Lead Contact", "Location", "Notes"]

    out_rows = []
    skipped = 0

    for row in rows:
        # Skip completely empty rows
        if not any(str(v).strip() for v in row.values() if v):
            continue

        start_raw = get(row, "Date")
        end_raw   = get(row, "End Date")
        date_field = format_date_field(start_raw, end_raw)

        # Skip TBD or unparseable dates
        if not date_field:
            skipped += 1
            continue

        event_name = get(row, "Event Name")
        if not event_name:
            continue

        status   = normalize_status(get(row, "Status"))
        team     = extract_team(event_name)
        audience = get(row, "Audience")
        lead     = get(row, "Lead Contact", "Lead")
        location = get(row, "Location")
        notes    = get(row, "Notes")

        out_rows.append({
            "Date":         date_field,
            "Status":       status,
            "Team":         team,
            "Event Name":   event_name,
            "Audience":     audience,
            "Lead Contact": lead,
            "Location":     location,
            "Notes":        notes,
        })

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=OUT_FIELDS)
        writer.writeheader()
        writer.writerows(out_rows)

    print(f"✅ Converted {len(out_rows)} events → {output_path}")
    if skipped:
        print(f"   ({skipped} TBD/unparseable rows skipped)")


if __name__ == "__main__":
    input_path  = sys.argv[1] if len(sys.argv) > 1 else "26-27_Admin_Calendar_Events_CTE-LL_Version.csv"
    output_path = sys.argv[2] if len(sys.argv) > 2 else "CTE-LL_Calendar_Events_2026-27.csv"

    if not os.path.isfile(input_path):
        print(f"❌ Input file not found: {input_path}")
        sys.exit(1)

    convert(input_path, output_path)
